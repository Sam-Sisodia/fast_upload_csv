from openpyxl import Workbook


def create_csv():
   
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Data"
    

    headers = ["email", "password", "phone"]
    for col,va in enumerate(headers,start=1):
        cell =ws.cell(row=1, column=col, value=va)


    row_index = 2
    for i in range(2,500000):
        email = f"admin@{i}.yopmail.com"
        print("++++++++++++++++",email)
        password = i
        phone = i*i
        cell =ws.cell(row=i, column=1, value=email)

        cell =ws.cell(row=i, column=2, value=password)
        cell =ws.cell(row=i, column=3, value=phone)


    wb.save("./5lakh.xlsx")
    



create_csv()


# import pandas as pd

# file_path = "10.xlsx"

# df = pd.read_excel(file_path)


# # for index, row in df.iterrows():
# #     # 'index' is the index of the row in the DataFrame
# #     # 'row' is a Pandas Series representing the row
# #     print(f"Row index: {index}")
# #     print(f"Row data:\n{row}\n")

# # print(df.iloc[0])

# # print("the mail is {email}")
# # print("the mail is {phone }")
# # print("the mail is {passwoed}" , coumn numer)

# for index, row in df.iterrows():
#     email = df.iloc[index]['email']  # Assuming 'email' is the column name
#     phone = df.iloc[index]['phone']  # Assuming 'phone' is the column name
#     password = df.iloc[index]['password']
#     print(f"the mail is {email}")
#     print(f"the mail is {phone }")
#     print(f"the mail is {password} numer")




# header  = []

# while True:
#     text = input("Enter the header name :")
#     if text =="":
#         break
#     else:
#         header.append(text)

# print("You entered:", header)
















# import pandas as pd


# file_path = "zz.xlsx"
# df = pd.read_excel(file_path)
# email_obj = df["email"]
# print(list(email_obj))


# def make_chunk_email(mail_list, chunk_size):
#     li = []

#     for c in range(0,len(mail_list),chunk_size):
#         chuck_list = mail_list[c: c+chunk_size]
#         li.append(chuck_list)

#         # print(chuck_list)
#     return li

# chunks_list = make_chunk_email(email_obj,2)

# for  chunk in chunks_list:
#     pass
    # print("+++++++++++++++++++++++++STRAT")
    # print(chunk)
    # print("___________________END")
    # print()









# # main.py
# from typing import List
# from fastapi import FastAPI, Depends, HTTPException
# from sqlalchemy.orm import Session
# from pydantic import BaseModel
# from models import Base, Item, SessionLocal, engine
# from sqlalchemy.exc import SQLAlchemyError

# app = FastAPI()

# # Dependency to get DB session
# def get_db():
#     db = SessionLocal()
#     try:
#         yield db
#     finally:
#         db.close()

# class ItemCreate(BaseModel):
#     name: str
#     description: str

# BATCH_SIZE = 1000  # Adjust the batch size as needed

# @app.post("/items/bulk/")
# def create_items(items: List[ItemCreate], db: Session = Depends(get_db)):
#     try:
#         for i in range(0, len(items), BATCH_SIZE):
#             batch = items[i:i + BATCH_SIZE]
#             db_items = [Item(name=item.name, description=item.description) for item in batch]
#             db.bulk_save_objects(db_items)
#             db.commit()
#         return {"status": "success", "inserted": len(items)}
#     except SQLAlchemyError as e:
#         db.rollback()
#         raise HTTPException(status_code=500, detail=str(e))





class A:
    e = 10

    def h(slef):
        print("+++++++++")



obj= A()
obj.e = 19
print(obj.e)
print(A.e)